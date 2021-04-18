#codeing=utf-8
import os
import sys
import openpyxl
base_path=os.getcwd()
sys.path.append(base_path)
# open_excel=openpyxl.load_workbook(base_path+"/Case/android4.0.1.xlsx")
# sheet_name=open_excel.sheetnames
# excel_value=open_excel[sheet_name[1]]
# print (excel_value)
# print (excel_value.cell(1,3).value)
# print (excel_value.max_row)

class HandExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        open_excel=openpyxl.load_workbook(base_path+"/Case/case.xlsx")
        return open_excel
    def get_sheet_data(self,index=None):
        '''
        加载所有sheet的内容
        '''
        sheet_name = self.load_excel().sheetnames
        if index == None:

            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data
    
    def get_cell_value(self,row,col):
        '''
        获取某一个单元格的内容
    
        '''
        data = self.get_sheet_data().cell(row=row,column=col).value
        return data
    
    def get_rows(self):
        '''
        获取行数
        '''
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self,row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_sheet_data()[row]:

            row_list.append(i.value)
        return row_list
    
    def excel_write_data(self,row,cols,data):
        '''
        往excel表格里写入数据
        '''
        wb=self.load_excel()
        '''
        激活excel表格
        '''
        wr=wb.active
        wr.cell(row,cols,data)
        wb.save(base_path+"/Case/case.xlsx")
    
    def get_colsa_value(self,col):
        #k=dict(self.get_sheet_data().iter_cols(10,10))
        col_list=[]
        cols=self.get_sheet_data().iter_cols(col,col)

        for col in cols:
            #print(col)
            for cell in col:
                #print(cell.value)
                col_list.append(cell.value)
        return col_list

            
        
        


handle_excel = HandExcel()
if __name__ == "__main__":
    handle_excel = HandExcel()
    print(handle_excel.load_excel())
    print('===========================================================')
    print(handle_excel.get_sheet_data())
    print('===========================================================')
    print(handle_excel.get_cell_value(3,2))
    print('===========================================================')
    print(handle_excel.get_rows_value(2))
    print(handle_excel.get_rows())
    # handle_excel.excel_write_data(2,9,"通过")
    # print(handle_excel.get_cols_value(10))
